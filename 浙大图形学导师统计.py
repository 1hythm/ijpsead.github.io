"""
浙江大学计算机图形学方向导师统计 — 生成 Excel 文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 样式定义 ──────────────────────────────────────
title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
section_font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
body_font = Font(name="微软雅黑", size=10)
small_font = Font(name="微软雅黑", size=9, color="666666")
link_font = Font(name="微软雅黑", size=9, color="0563C1", underline="single")

title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# ── 数据 ──────────────────────────────────────────
professors = [
    # (序号, 姓名, 职称/荣誉, 研究方向, 所属团队, 所属学院, 备注)
    ("教授 / 长聘教授（博导）", [
        (1, "周昆", "教授，长江学者，国家杰青，IEEE/ACM Fellow", "计算机图形学、人机交互、VR、并行计算", "GAPS（负责人）", "计算机学院", "CAD&CG实验室主任"),
        (2, "鲍虎军", "教授，长江学者，国家杰青", "计算机图形学、几何计算、VR、视觉计算", "混合现实（负责人）", "计算机学院", ""),
        (3, "彭群生", "教授", "图形学基础理论、真实感绘制、动画、科学可视化", "CAD&CG实验室", "计算机学院", "学术委员会副主任"),
        (4, "金小刚", "二级教授", "计算机动画、VR、数字化服装/医疗、交通仿真", "CAD&CG实验室", "计算机学院", "省VR联盟理事长"),
        (5, "许威威", "长聘教授，长江学者", "三维感知/重建/仿真、数字孪生、VR", "混合现实", "计算机学院", ""),
        (6, "黄劲", "教授，国家优青", "几何计算、物理模拟、数字几何处理", "混合现实", "计算机学院", ""),
        (7, "吴鸿智", "教授，国家优青", "外观建模、高维材质、文物数字化保护", "GAPS", "计算机学院", ""),
        (8, "童若锋", "教授", "图形学、CAD、视觉、视频分析、医学图像", "CAD&CG实验室", "计算机学院", "实验室副主任"),
        (9, "刘新国", "教授，博导", "图形学、VR、视觉、人体建模、几何处理", "CAD&CG实验室", "计算机学院", ""),
        (10, "冯结青", "教授，博导", "CAD、图形学、代数曲面建模与绘制", "CAD&CG实验室", "计算机学院", ""),
        (11, "王锐", "教授", "实时/真实感渲染、GPU计算、3D显示", "CAD&CG实验室", "计算机学院", ""),
        (12, "于金辉", "研究员，博导", "计算机动画、文化遗产数字化保护", "CAD&CG实验室", "计算机学院", ""),
    ]),
    ("副教授 / 青年研究员（硕导/博导）", [
        (13, "任重", "副教授，博导", "真实感/实时绘制、阴影/体绘制、VR", "GAPS", "计算机学院", "考研报考联系人"),
        (14, "张宏鑫", "副教授", "CAGD、图形学、云计算、大数据可视化", "CAD&CG实验室", "计算机学院", "计算机系副主任"),
        (15, "李明", "副教授", "CAD/CAE一体化、智能设计与仿真", "—", "计算机学院", "软件工程系副主任"),
        (16, "侯启明", "副教授", "计算机图形学、并行计算、编译器", "GAPS", "计算机学院", ""),
        (17, "陈翔", "副教授", "计算机图形学、3D打印", "GAPS", "计算机学院", ""),
        (18, "邵天甲", "百人计划研究员，博导，国家优青", "图形学、3D视觉、3D AIGC、数字人", "GAPS", "计算机学院", ""),
        (19, "彭思达", "百人计划研究员，博导", "三维视觉、图形学（重建/动态场景/3DGS）", "混合现实", "软件学院", "China3DV 2025杰出青年学者"),
        (20, "翁荻", "百人计划研究员，博导", "时空大数据交互式治理与可视分析", "—", "软件学院", ""),
        (21, "郑友怡", "研究员", "计算机图形学、图像视频处理、人机交互", "GAPS", "计算机学院", ""),
        (22, "楼建文", "特聘研究员", "图形学、VR、多媒体（几何/表观建模与渲染）", "GAPS", "软件学院", "软件学院夏令营联系人"),
        (23, "王驰", "特聘研究员", "AIGC、图形学（三维生成、多模态大模型）", "—", "软件学院", ""),
        (24, "邹常青", "研究员", "图形学、计算机视觉、AIGC", "混合现实", "计算机学院", ""),
        (25, "丁尧相", "研究员", "机器学习理论与算法", "GAPS", "计算机学院", ""),
        (26, "冯天", "特聘副研究员 ⚠️2026起不再招生", "图形学、VR、人机交互、智能设计", "GAPS", "—", ""),
    ]),
    ("跨院系相关导师", [
        (27, "谭建荣", "教授，中国工程院院士", "CAD&CG、数字化设计与制造", "—", "机械工程学院", ""),
        (28, "李基拓", "副教授", "CAD&CG、3D人体/服装建模、仿人机器人", "—", "机械工程学院", ""),
    ]),
]

# ── Sheet 1: 导师总表 ─────────────────────────────
ws = wb.active
ws.title = "导师总表"

# 列宽
col_widths = [6, 10, 32, 42, 18, 16, 26]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 标题行
ws.merge_cells("A1:G1")
ws["A1"] = "浙江大学 计算机图形学方向导师统计"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws["A1"].alignment = center_align
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:G2")
ws["A2"] = "数据来源：公开搜索整理 | 统计日期：2026年7月 | 核心基地：CAD&CG全国重点实验室"
ws["A2"].font = small_font
ws["A2"].alignment = center_align
ws.row_dimensions[2].height = 22

headers = ["序号", "姓名", "职称/荣誉", "研究方向", "所属团队", "所属学院", "备注"]
row_idx = 4
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=row_idx, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[row_idx].height = 28

row_idx = 5
for section_title, group in professors:
    # 分组标题行
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    ws.cell(row=row_idx, column=1, value=section_title).font = section_font
    ws.cell(row=row_idx, column=1).fill = section_fill
    ws.cell(row=row_idx, column=1).alignment = left_align
    for c in range(1, 8):
        ws.cell(row=row_idx, column=c).border = thin_border
        ws.cell(row=row_idx, column=c).fill = section_fill
    ws.row_dimensions[row_idx].height = 26
    row_idx += 1

    for item in group:
        for col_idx, val in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if col_idx in (1, 2, 6):
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            # 斑马条纹
            if item[0] % 2 == 0:
                cell.fill = even_fill
        ws.row_dimensions[row_idx].height = 36
        row_idx += 1

# 底部统计
row_idx += 1
ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
ws.cell(row=row_idx, column=1, value="📊 统计：共计 28 人 ｜ 教授级 12 人 ｜ 副教授/青年研究员 14 人 ｜ 跨院系 2 人 ｜ 国家级人才（杰青/长江/优青/院士）约 8 人").font = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
ws.cell(row=row_idx, column=1).fill = yellow_fill
ws.cell(row=row_idx, column=1).alignment = center_align
for c in range(1, 8):
    ws.cell(row=row_idx, column=c).border = thin_border

# ── Sheet 2: 团队速览 ─────────────────────────────
ws2 = wb.create_sheet("团队速览")

ws2.merge_cells("A1:E1")
ws2["A1"] = "浙江大学 CAD&CG 两大核心团队速览"
ws2["A1"].font = title_font
ws2["A1"].fill = title_fill
ws2["A1"].alignment = center_align
ws2.row_dimensions[1].height = 40

teams = [
    ("GAPS（图形与并行系统）", "周昆", "任重、吴鸿智、侯启明、陈翔、郑友怡、邵天甲、丁尧相、楼建文、冯天",
     "计算机图形学 + 并行计算、外观建模、3D打印、3D AIGC",
     "renzhong@zju.edu.cn（任重）"),
    ("混合现实", "鲍虎军", "黄劲、许威威、王锐、邹常青、彭思达",
     "VR/AR、几何计算、三维感知与重建",
     "pengsida@zju.edu.cn（彭思达）"),
]

team_headers = ["团队名称", "负责人", "核心成员", "特色方向", "联系方式"]
for col_idx, h in enumerate(team_headers, 1):
    cell = ws2.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for t_idx, team in enumerate(teams):
    for col_idx, val in enumerate(team, 1):
        cell = ws2.cell(row=4 + t_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = left_align if col_idx > 1 else center_align
    ws2.row_dimensions[4 + t_idx].height = 50

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 55
ws2.column_dimensions["D"].width = 45
ws2.column_dimensions["E"].width = 30

# ── Sheet 3: 招生提示 ─────────────────────────────
ws3 = wb.create_sheet("报考建议")

tips = [
    "报考浙江大学图形学方向研究生的建议",
    "",
    "1. 计算机学院 vs 软件学院",
    "   • 计算机学院：导师数量多（约22人），研究方向全面，CAD&CG实验室为主体",
    "   • 软件学院：彭思达、翁荻、楼建文、王驰等在软件学院直接任职",
    "   • 部分导师可在两个学院同时招生，需查看当年简章或直接询问导师",
    "",
    "2. 如何联系导师",
    "   • 访问浙江大学教师主页：https://person.zju.edu.cn/",
    "   • CAD&CG实验室官网：http://www.cad.zju.edu.cn",
    "   • 计算机学院官网：http://www.cs.zju.edu.cn/",
    "   • 软件学院官网：http://www.cst.zju.edu.cn/",
    "   • 建议提前发送邮件（附简历+成绩单），表达研究兴趣",
    "",
    "3. GAPS团队考研联系人",
    "   • 任重老师：renzhong@zju.edu.cn",
    "   • 楼建文老师：jianwen.lou@zju.edu.cn（软件学院夏令营）",
    "",
    "4. 热门方向（供参考）",
    "   • 三维重建 & 3D AIGC：邵天甲、彭思达、王驰、邹常青",
    "   • 渲染 & 外观建模：吴鸿智、任重、王锐",
    "   • VR/AR：鲍虎军、许威威、金小刚",
    "   • 几何处理 & 物理模拟：黄劲、刘新国",
    "   • 可视化 & 数据分析：翁荻、张宏鑫",
    "",
    "⚠️ 以上信息基于2026年7月公开搜索，每年招生情况可能变化，务必确认最新信息。",
]

for i, line in enumerate(tips, 1):
    cell = ws3.cell(row=i, column=1, value=line)
    if line.startswith("报考") or line == "":
        continue
    if line.startswith("⚠️"):
        cell.font = Font(name="微软雅黑", size=10, color="CC0000", bold=True)
    elif line.startswith("1.") or line.startswith("2.") or line.startswith("3.") or line.startswith("4."):
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
    elif line.startswith("   •"):
        cell.font = body_font
    else:
        cell.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")

ws3.column_dimensions["A"].width = 80

# ── 保存 ──────────────────────────────────────────
output_path = r"C:\Users\22160\Desktop\work\浙大图形学导师统计.xlsx"
wb.save(output_path)
print(f"Excel saved: {output_path}")
